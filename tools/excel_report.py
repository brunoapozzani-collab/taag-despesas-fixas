"""Excel executive-level report — 6-tab package for CEO decision-making."""
from __future__ import annotations

import io
from datetime import date

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from expense_engine import COMPANIES, assign_ceo_category, monthly_by_company, monthly_total

# ---- Colour palette ----
CYAN_HEX   = "34B3D3"
BLACK_HEX  = "000000"
DGREY_HEX  = "444444"
LIGHT_HEX  = "F3F3F3"
LBLUE_HEX  = "E8F7FC"
AMBER_HEX  = "FFF3CD"
WHITE_HEX  = "FFFFFF"

HEADER_FILL  = PatternFill("solid", fgColor=CYAN_HEX)
BLACK_FILL   = PatternFill("solid", fgColor=BLACK_HEX)
LIGHT_FILL   = PatternFill("solid", fgColor=LIGHT_HEX)
LBLUE_FILL   = PatternFill("solid", fgColor=LBLUE_HEX)
AMBER_FILL   = PatternFill("solid", fgColor=AMBER_HEX)

HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color=WHITE_HEX)
TITLE_FONT   = Font(name="Calibri", size=15, bold=True, color=WHITE_HEX)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE_HEX)
BODY_FONT    = Font(name="Calibri", size=10, color=BLACK_HEX)
BODY_BOLD    = Font(name="Calibri", size=10, bold=True, color=BLACK_HEX)
SMALL_FONT   = Font(name="Calibri", size=9,  color="888888")

THIN   = Side(style="thin",   color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PALETTE = ["#34b3d3", "#1d6e85", "#000000", "#666666", "#a8e1ee", "#cccccc"]

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


def brl(x: float) -> str:
    s = f"R$ {x:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _autosize(ws, min_w: int = 8, max_w: int = 52):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        mlen = min_w
        for cell in col:
            if cell.value is not None:
                mlen = max(mlen, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[cl].width = mlen


def _sec_header(ws, row: int, col: int, text: str, n_cols: int = 1, fill=None):
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + n_cols - 1)
    c = ws.cell(row=row, column=col, value=f"  {text}")
    c.font = SECTION_FONT
    c.fill = fill or HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def _col_headers(ws, row: int, headers: list[str], start_col: int = 1, fill=None):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = fill or HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


# ---- Insight generator (programmatic, no API) ----

def _insights(df_fixed: pd.DataFrame, monthly: pd.DataFrame) -> list[str]:
    grand_total = float(df_fixed["Valor"].abs().sum())
    if grand_total == 0:
        return []
    items = []

    by_co = (df_fixed.assign(V=df_fixed["Valor"].abs())
             .groupby("Empresa", as_index=False)["V"].sum()
             .rename(columns={"V": "Total"})
             .sort_values("Total", ascending=False))
    if not by_co.empty:
        top = by_co.iloc[0]
        pct = top["Total"] / grand_total * 100
        items.append(f"📍  Localidade mais cara: {top['Empresa']} — {brl(top['Total'])} ({pct:.1f}% do total)")

    by_ceo = (df_fixed.assign(V=df_fixed["Valor"].abs())
              .groupby("CeoCategoria", as_index=False)["V"].sum()
              .rename(columns={"V": "Total"})
              .sort_values("Total", ascending=False))
    if not by_ceo.empty:
        top_c = by_ceo.iloc[0]
        pct_c = top_c["Total"] / grand_total * 100
        items.append(f"📊  Maior categoria: {top_c['CeoCategoria']} — {brl(top_c['Total'])} ({pct_c:.1f}% do total)")

    alg = by_ceo[by_ceo["CeoCategoria"] == "Aluguel"]
    if not alg.empty:
        alg_pct = float(alg["Total"].sum()) / grand_total * 100
        if alg_pct > 35:
            items.append(f"⚠️   Aluguel consome {alg_pct:.1f}% das despesas fixas — renegociação pode gerar economia significativa")
        else:
            items.append(f"🏢  Aluguel representa {alg_pct:.1f}% das despesas fixas totais")

    if len(monthly) >= 2:
        last  = float(monthly["Total"].iloc[-1])
        prev  = float(monthly["Total"].iloc[-2])
        delta = (last - prev) / prev * 100 if prev else 0
        mes   = str(monthly["Mes"].iloc[-1])
        arrow = "▲" if delta > 3 else ("▼" if delta < -3 else "→")
        items.append(f"{arrow}  Último mês ({mes}): {brl(last)}  ({delta:+.1f}% vs. mês anterior)")

    n_outros = int((df_fixed["Empresa"] == "Outros").sum())
    if n_outros > 0:
        val_outros = float(df_fixed[df_fixed["Empresa"] == "Outros"]["Valor"].abs().sum())
        items.append(f"🔍  {n_outros} lançamentos sem localidade ({brl(val_outros)}) — requerem atribuição")

    of = by_ceo[by_ceo["CeoCategoria"] == "Outros Fixos"]
    if not of.empty:
        items.append(f"📋  'Outros Fixos': {brl(float(of['Total'].sum()))} — despesas fora das categorias padrão")

    return items


# ---- Tab 1: Executive Summary ----

def _write_tab1(ws, df_fixed: pd.DataFrame, start: date, end: date):
    period    = f"Período: {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"
    today_str = date.today().strftime("%d/%m/%Y")

    # Title bar
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="TAAG BRASIL — RELATÓRIO EXECUTIVO DE DESPESAS FIXAS")
    t.font = TITLE_FONT; t.fill = BLACK_FILL; t.alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"{period}    |    Gerado em {today_str}    |    CONFIDENCIAL").font = SMALL_FONT
    ws.row_dimensions[2].height = 14

    grand_total = float(df_fixed["Valor"].abs().sum())
    n_co   = df_fixed["Empresa"].nunique()
    n_rows = len(df_fixed)
    by_co  = (df_fixed.assign(V=df_fixed["Valor"].abs())
              .groupby("Empresa", as_index=False)["V"].sum()
              .rename(columns={"V": "Total"}).sort_values("Total", ascending=False))
    top_co = str(by_co.iloc[0]["Empresa"]) if not by_co.empty else "—"

    # KPI block (row 4 = value, row 5 = label)
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 16
    for i, (label, val) in enumerate([
        ("TOTAL DESPESAS FIXAS", brl(grand_total)),
        ("LOCALIDADES",          str(n_co)),
        ("LANÇAMENTOS",          str(n_rows)),
        ("MAIOR LOCALIDADE",     top_co),
    ]):
        c = i * 2 + 1
        for row_i in (4, 5):
            ws.merge_cells(start_row=row_i, start_column=c, end_row=row_i, end_column=c + 1)
        v = ws.cell(row=4, column=c, value=val)
        v.font = Font(name="Calibri", size=15, bold=True, color=WHITE_HEX)
        v.fill = HEADER_FILL; v.alignment = CENTER
        l = ws.cell(row=5, column=c, value=label)
        l.font = Font(name="Calibri", size=8, color="888888"); l.alignment = CENTER

    # Top Categories
    r = 7
    _sec_header(ws, r, 1, "TOP CATEGORIAS DE DESPESA", n_cols=4)
    r += 1
    _col_headers(ws, r, ["Categoria", "Total (R$)", "% do Total", "Ação Sugerida"])
    r += 1
    by_ceo = (df_fixed.assign(V=df_fixed["Valor"].abs())
              .groupby("CeoCategoria", as_index=False)["V"].sum()
              .rename(columns={"V": "Total"}).sort_values("Total", ascending=False))
    for idx, row in by_ceo.head(10).iterrows():
        pct    = row["Total"] / grand_total * 100 if grand_total else 0
        action = ("⚠️ Renegociar" if pct > 40 and "Aluguel" in str(row["CeoCategoria"])
                  else ("Monitorar" if pct > 8 else "OK"))
        ws.cell(row=r, column=1, value=str(row["CeoCategoria"])).font = BODY_FONT
        c2 = ws.cell(row=r, column=2, value=float(row["Total"]))
        c2.font = BODY_FONT; c2.number_format = 'R$ #,##0.00'; c2.alignment = RIGHT
        c3 = ws.cell(row=r, column=3, value=round(pct / 100, 4))
        c3.font = BODY_FONT; c3.number_format = "0.0%"; c3.alignment = RIGHT
        ws.cell(row=r, column=4, value=action).font = BODY_FONT
        if r % 2 == 0:
            for ci in range(1, 5):
                ws.cell(row=r, column=ci).fill = LIGHT_FILL
        r += 1

    # Top Locations
    r += 1
    _sec_header(ws, r, 1, "DESEMPENHO POR LOCALIDADE", n_cols=4)
    r += 1
    _col_headers(ws, r, ["Localidade", "Total (R$)", "% do Total", "Vs. Média"])
    r += 1
    avg = grand_total / n_co if n_co else 0
    for _, row in by_co.iterrows():
        val = float(row["Total"]); pct = val / grand_total * 100 if grand_total else 0
        ws.cell(row=r, column=1, value=str(row["Empresa"])).font = BODY_FONT
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = BODY_FONT; c2.number_format = 'R$ #,##0.00'; c2.alignment = RIGHT
        c3 = ws.cell(row=r, column=3, value=round(pct / 100, 4))
        c3.font = BODY_FONT; c3.number_format = "0.0%"; c3.alignment = RIGHT
        delta = val - avg
        c4 = ws.cell(row=r, column=4, value=delta)
        c4.font = Font(name="Calibri", size=10, color=("006400" if delta < 0 else "B22222"))
        c4.number_format = 'R$ #,##0.00'; c4.alignment = RIGHT
        r += 1

    # Strategic Insights
    r += 1
    _sec_header(ws, r, 1, "DESTAQUES E RECOMENDAÇÕES ESTRATÉGICAS", n_cols=6)
    r += 1
    monthly = monthly_total(df_fixed)
    for insight in _insights(df_fixed, monthly):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=insight)
        c.font = BODY_FONT; c.fill = LBLUE_FILL; c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 18
        r += 1

    ws.freeze_panes = "A3"
    _autosize(ws)


# ---- Tab 2: Monthly Comparison ----

def _write_tab2(ws, df_fixed: pd.DataFrame, period: str):
    ws.cell(row=1, column=1, value="EVOLUÇÃO MENSAL POR LOCALIDADE").font = BODY_BOLD
    ws.cell(row=2, column=1, value=period).font = SMALL_FONT

    pivot = monthly_by_company(df_fixed)
    if pivot.empty:
        ws.cell(row=4, column=1, value="Sem dados no período.").font = BODY_FONT
        return

    co_cols = list(pivot.columns)
    _col_headers(ws, 4, ["Mês"] + co_cols + ["TOTAL", "Variação"])

    r = 5; prev_total = None
    for mes, row in pivot.iterrows():
        total = float(row.sum())
        ws.cell(row=r, column=1, value=str(mes)).font = BODY_FONT
        for i, c in enumerate(co_cols, start=2):
            cell = ws.cell(row=r, column=i, value=float(row[c]))
            cell.font = BODY_FONT; cell.number_format = 'R$ #,##0.00'
            cell.border = BORDER; cell.alignment = RIGHT
        tc = ws.cell(row=r, column=len(co_cols) + 2, value=total)
        tc.font = BODY_BOLD; tc.number_format = 'R$ #,##0.00'; tc.fill = LIGHT_FILL; tc.alignment = RIGHT
        if prev_total and prev_total > 0:
            pct = (total - prev_total) / prev_total * 100
            vc = ws.cell(row=r, column=len(co_cols) + 3, value=f"{pct:+.1f}%")
            vc.font = Font(name="Calibri", size=10,
                           color=("B22222" if pct > 5 else ("006400" if pct < -5 else DGREY_HEX)))
        prev_total = total
        r += 1

    # Grand total row
    ws.cell(row=r, column=1, value="TOTAL GERAL").font = BODY_BOLD
    for i, c in enumerate(co_cols, start=2):
        cell = ws.cell(row=r, column=i, value=float(pivot[c].sum()))
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE_HEX)
        cell.fill = BLACK_FILL; cell.number_format = 'R$ #,##0.00'; cell.alignment = RIGHT
    gt = ws.cell(row=r, column=len(co_cols) + 2, value=float(pivot.values.sum()))
    gt.font = Font(name="Calibri", size=10, bold=True, color=WHITE_HEX)
    gt.fill = BLACK_FILL; gt.number_format = 'R$ #,##0.00'; gt.alignment = RIGHT

    ws.freeze_panes = "B5"
    _autosize(ws)


# ---- Tab 3: Location Comparison ----

def _write_tab3(ws, df_fixed: pd.DataFrame):
    ws.cell(row=1, column=1, value="COMPARATIVO POR CATEGORIA × LOCALIDADE").font = BODY_BOLD
    ws.cell(row=2, column=1, value="Destaque azul = maior gasto nessa categoria").font = SMALL_FONT

    pivot = (df_fixed.assign(V=df_fixed["Valor"].abs())
             .pivot_table(index="CeoCategoria", columns="Empresa",
                          values="V", aggfunc="sum", fill_value=0))
    if pivot.empty:
        ws.cell(row=4, column=1, value="Sem dados.").font = BODY_FONT; return

    co_cols = [c for c in COMPANIES if c in pivot.columns] + \
              [c for c in pivot.columns if c not in COMPANIES]
    pivot = pivot[[c for c in co_cols if c in pivot.columns]]
    pivot["TOTAL"] = pivot.sum(axis=1)
    grand = float(pivot["TOTAL"].sum())
    pivot = pivot.sort_values("TOTAL", ascending=False)

    headers = ["Categoria"] + [c for c in co_cols if c in pivot.columns] + ["TOTAL", "% DO TOTAL"]
    _col_headers(ws, 4, headers)

    r = 5
    for cat, row in pivot.iterrows():
        ws.cell(row=r, column=1, value=str(cat)).font = BODY_FONT
        vals = [float(row.get(c, 0)) for c in co_cols if c in pivot.columns]
        max_v = max(vals) if vals else 0
        for i, c in enumerate([c for c in co_cols if c in pivot.columns], start=2):
            v = float(row.get(c, 0))
            cell = ws.cell(row=r, column=i, value=v if v > 0 else None)
            cell.number_format = 'R$ #,##0.00'; cell.border = BORDER; cell.alignment = RIGHT
            if v > 0 and v == max_v:
                cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE_HEX)
                cell.fill = HEADER_FILL
            else:
                cell.font = BODY_FONT
                if r % 2 == 0:
                    cell.fill = LIGHT_FILL
        n_data_cols = len([c for c in co_cols if c in pivot.columns])
        tc = ws.cell(row=r, column=n_data_cols + 2, value=float(row["TOTAL"]))
        tc.font = BODY_BOLD; tc.number_format = 'R$ #,##0.00'; tc.fill = LIGHT_FILL; tc.alignment = RIGHT
        pct = float(row["TOTAL"]) / grand * 100 if grand else 0
        pc = ws.cell(row=r, column=n_data_cols + 3, value=round(pct / 100, 4))
        pc.font = BODY_FONT; pc.number_format = "0.0%"; pc.alignment = RIGHT
        r += 1

    # Grand total row
    n_dc = len([c for c in co_cols if c in pivot.columns])
    ws.cell(row=r, column=1, value="TOTAL GERAL").font = Font(name="Calibri", size=10, bold=True, color=WHITE_HEX)
    ws.cell(row=r, column=1).fill = BLACK_FILL
    for i, c in enumerate([c for c in co_cols if c in pivot.columns], start=2):
        cell = ws.cell(row=r, column=i, value=float(pivot[c].sum()))
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE_HEX)
        cell.fill = BLACK_FILL; cell.number_format = 'R$ #,##0.00'; cell.alignment = RIGHT
    gt = ws.cell(row=r, column=n_dc + 2, value=grand)
    gt.font = Font(name="Calibri", size=10, bold=True, color=WHITE_HEX)
    gt.fill = BLACK_FILL; gt.number_format = 'R$ #,##0.00'; gt.alignment = RIGHT

    ws.freeze_panes = "B5"
    _autosize(ws)


# ---- Tab 4: Detailed by Category ----

def _write_tab4(ws, df_fixed: pd.DataFrame, period: str):
    ws.cell(row=1, column=1, value="DESPESAS FIXAS DETALHADAS POR CATEGORIA").font = BODY_BOLD
    ws.cell(row=2, column=1, value=period).font = SMALL_FONT

    COLS = [
        ("Pagto",       "Data"),
        ("Empresa",     "Localidade"),
        ("Favorecido",  "Fornecedor"),
        ("Descricao",   "Descrição"),
        ("Despesas",    "Conta"),
        ("Valor",       "Valor (R$)"),
    ]
    r = 4
    for cat in sorted(df_fixed["CeoCategoria"].dropna().unique()):
        sub = df_fixed[df_fixed["CeoCategoria"] == cat].sort_values(["Empresa", "Pagto"])
        if sub.empty:
            continue

        # Category header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        c = ws.cell(row=r, column=1, value=f"  {cat.upper()}")
        c.font = SECTION_FONT; c.fill = HEADER_FILL
        ws.row_dimensions[r].height = 18
        r += 1

        # Column headers
        for i, (_, h) in enumerate(COLS, start=1):
            hc = ws.cell(row=r, column=i, value=h)
            hc.font = Font(name="Calibri", size=9, bold=True); hc.fill = LIGHT_FILL; hc.border = BORDER
        r += 1

        # Rows
        for _, row in sub.iterrows():
            for i, (key, _) in enumerate(COLS, start=1):
                val = row.get(key)
                if key == "Pagto" and pd.notna(val):
                    val = pd.to_datetime(val).strftime("%d/%m/%Y")
                elif key == "Valor" and pd.notna(val):
                    val = float(abs(val))
                cell = ws.cell(row=r, column=i, value=val)
                cell.font = BODY_FONT; cell.border = BORDER
                if key == "Valor":
                    cell.number_format = 'R$ #,##0.00'; cell.alignment = RIGHT
            r += 1

        # Subtotal
        sub_total = float(sub["Valor"].abs().sum())
        ws.cell(row=r, column=len(COLS) - 1, value="Subtotal:").font = BODY_BOLD
        sc = ws.cell(row=r, column=len(COLS), value=sub_total)
        sc.font = Font(name="Calibri", size=10, bold=True, color=WHITE_HEX)
        sc.fill = HEADER_FILL; sc.number_format = 'R$ #,##0.00'; sc.alignment = RIGHT
        r += 2

    ws.freeze_panes = "A5"
    _autosize(ws, min_w=10)


# ---- Tab 5: Anomalies ----

def _write_tab5(ws, df_fixed: pd.DataFrame, df_excluded: pd.DataFrame, period: str):
    ws.cell(row=1, column=1, value="ANOMALIAS E PONTOS DE ATENÇÃO").font = BODY_BOLD
    ws.cell(row=2, column=1, value=period).font = SMALL_FONT

    ACOLS = ["Data", "Localidade", "Fornecedor", "Descrição", "Categoria CEO", "Valor"]
    NA = "—"

    def _write_rows(ws, r, subset):
        _col_headers(ws, r, ACOLS, fill=PatternFill("solid", fgColor="888888"))
        r += 1
        for _, row in subset.iterrows():
            pagto = pd.to_datetime(row.get("Pagto"))
            ws.cell(row=r, column=1, value=pagto.strftime("%d/%m/%Y") if pd.notna(pagto) else NA).font = BODY_FONT
            ws.cell(row=r, column=2, value=str(row.get("Empresa", NA))).font = BODY_FONT
            ws.cell(row=r, column=3, value=str(row.get("Favorecido", NA))[:55]).font = BODY_FONT
            ws.cell(row=r, column=4, value=str(row.get("Descricao", NA))[:55]).font = BODY_FONT
            ws.cell(row=r, column=5, value=str(row.get("CeoCategoria", NA))).font = BODY_FONT
            vc = ws.cell(row=r, column=6, value=float(abs(row.get("Valor", 0))))
            vc.font = BODY_FONT; vc.number_format = 'R$ #,##0.00'; vc.alignment = RIGHT
            for ci in range(1, 7):
                ws.cell(row=r, column=ci).fill = AMBER_FILL
            r += 1
        return r

    r = 4
    # 1. Unassigned location
    outros = df_fixed[df_fixed["Empresa"] == "Outros"]
    _sec_header(ws, r, 1, f"LANÇAMENTOS SEM LOCALIDADE DEFINIDA  ({len(outros)} linhas)",
                n_cols=6, fill=PatternFill("solid", fgColor="D97706"))
    r += 1
    if outros.empty:
        ws.cell(row=r, column=1, value="✓  Todos os lançamentos têm localidade definida.").font = BODY_FONT; r += 2
    else:
        r = _write_rows(ws, r, outros)
        r += 1

    # 2. Uncategorized
    of = df_fixed[df_fixed["CeoCategoria"] == "Outros Fixos"]
    _sec_header(ws, r, 1, f"DESPESAS NÃO MAPEADAS EM CATEGORIA PADRÃO  ({len(of)} linhas)",
                n_cols=6, fill=PatternFill("solid", fgColor="D97706"))
    r += 1
    if of.empty:
        ws.cell(row=r, column=1, value="✓  Todas as despesas estão mapeadas.").font = BODY_FONT; r += 2
    else:
        r = _write_rows(ws, r, of)
        r += 1

    # 3. Excluded expenses
    if not df_excluded.empty:
        _sec_header(ws, r, 1, f"DESPESAS EXCLUÍDAS PELO USUÁRIO  ({len(df_excluded)} linhas)",
                    n_cols=6, fill=PatternFill("solid", fgColor="888888"))
        r += 1
        r = _write_rows(ws, r, df_excluded.head(100))

    _autosize(ws, min_w=10)


# ---- Tab 6: Visual Dashboard ----

def _chart_buf(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _pie_categories(df_fixed: pd.DataFrame) -> io.BytesIO:
    by_ceo = (df_fixed.assign(V=df_fixed["Valor"].abs())
              .groupby("CeoCategoria")["V"].sum().sort_values(ascending=False))
    top = by_ceo.head(8)
    rest = by_ceo.iloc[8:].sum()
    if rest > 0:
        top["Outros"] = rest
    fig, ax = plt.subplots(figsize=(7, 5))
    wedges, _, autotexts = ax.pie(
        top.values, labels=None, colors=PALETTE * 4,
        autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
        startangle=140, wedgeprops=dict(edgecolor="white", linewidth=1.5),
    )
    for at in autotexts:
        at.set_fontsize(7)
    ax.legend(wedges, top.index.tolist(),
              loc="center left", bbox_to_anchor=(1.0, 0.5), frameon=False, fontsize=8)
    ax.set_title("Distribuição por Categoria", fontsize=10, fontweight="bold")
    fig.tight_layout()
    return _chart_buf(fig)


def _bar_locations(df_fixed: pd.DataFrame) -> io.BytesIO:
    by_co = (df_fixed.assign(V=df_fixed["Valor"].abs())
             .groupby("Empresa")["V"].sum().sort_values(ascending=False))
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.bar(by_co.index, by_co.values, color=PALETTE[:len(by_co)], edgecolor="black", lw=0.5)
    ax.set_title("Total por Localidade", fontsize=10, fontweight="bold")
    ax.set_ylabel("R$")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"R${x/1000:.0f}k" if x >= 1000 else f"R${x:.0f}"))
    ax.tick_params(axis="x", labelrotation=15, labelsize=8)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)
    fig.tight_layout()
    return _chart_buf(fig)


def _bar_stacked(df_fixed: pd.DataFrame) -> io.BytesIO:
    pivot = (df_fixed.assign(V=df_fixed["Valor"].abs())
             .pivot_table(index="Empresa", columns="CeoCategoria", values="V",
                          aggfunc="sum", fill_value=0))
    top_cats = (df_fixed.assign(V=df_fixed["Valor"].abs())
                .groupby("CeoCategoria")["V"].sum()
                .sort_values(ascending=False).head(7).index.tolist())
    cols = [c for c in top_cats if c in pivot.columns]
    pivot = pivot[cols]
    fig, ax = plt.subplots(figsize=(8, 5))
    bottom = [0] * len(pivot)
    for i, col in enumerate(cols):
        ax.bar(pivot.index, pivot[col], bottom=bottom,
               color=PALETTE[i % len(PALETTE)], label=col, edgecolor="white", lw=0.5)
        bottom = [b + v for b, v in zip(bottom, pivot[col])]
    ax.set_title("Composição por Localidade e Categoria", fontsize=10, fontweight="bold")
    ax.set_ylabel("R$")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"R${x/1000:.0f}k" if x >= 1000 else f"R${x:.0f}"))
    ax.tick_params(axis="x", labelrotation=15, labelsize=8)
    ax.legend(bbox_to_anchor=(1.02, 1.0), loc="upper left", frameon=False, fontsize=8)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)
    fig.tight_layout()
    return _chart_buf(fig)


def _line_trend(df_fixed: pd.DataFrame) -> io.BytesIO:
    monthly = monthly_total(df_fixed)
    fig, ax = plt.subplots(figsize=(8, 4))
    if not monthly.empty:
        x = range(len(monthly))
        ax.plot(monthly["Mes"], monthly["Total"], marker="o", lw=2.5,
                color="#34b3d3", markerfacecolor="#000000", markersize=7)
        ax.fill_between(x, monthly["Total"], alpha=0.12, color="#34b3d3")
        for i, (m, t) in enumerate(zip(monthly["Mes"], monthly["Total"])):
            ax.annotate(f"R${t/1000:.0f}k" if t >= 1000 else f"R${t:.0f}",
                        (i, t), textcoords="offset points", xytext=(0, 10),
                        ha="center", fontsize=8, fontweight="bold")
        ax.set_ylim(0, monthly["Total"].max() * 1.22)
        ax.set_title("Evolução Mensal das Despesas Fixas", fontsize=10, fontweight="bold")
        ax.set_ylabel("R$")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"R${x/1000:.0f}k" if x >= 1000 else f"R${x:.0f}"))
        ax.tick_params(axis="x", labelrotation=30, labelsize=9)
        for sp in ("top", "right"):
            ax.spines[sp].set_visible(False)
    fig.tight_layout()
    return _chart_buf(fig)


def _grouped_bar(df_fixed: pd.DataFrame) -> io.BytesIO:
    pivot = (df_fixed.assign(V=df_fixed["Valor"].abs())
             .pivot_table(index="CeoCategoria", columns="Empresa", values="V",
                          aggfunc="sum", fill_value=0))
    pivot["_t"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("_t", ascending=False).drop(columns="_t").head(10).iloc[::-1]
    companies = list(pivot.columns)
    n_cos = len(companies); bar_h = 0.7 / max(n_cos, 1)
    fig, ax = plt.subplots(figsize=(9, max(5, len(pivot) * 0.55 + 1.5)))
    for i, col in enumerate(companies):
        offset = (i - n_cos / 2 + 0.5) * bar_h
        ax.barh([p + offset for p in range(len(pivot))], pivot[col],
                height=bar_h * 0.88, color=PALETTE[i % len(PALETTE)], label=col)
    ax.set_yticks(list(range(len(pivot))))
    ax.set_yticklabels([str(c)[:30] for c in pivot.index], fontsize=8)
    ax.set_xlabel("R$")
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"R${x/1000:.0f}k" if x >= 1000 else f"R${x:.0f}"))
    ax.legend(bbox_to_anchor=(1.02, 1.0), loc="upper left", frameon=False, fontsize=8)
    ax.set_title("Top Categorias — Comparativo por Localidade", fontsize=10, fontweight="bold")
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)
    fig.tight_layout()
    return _chart_buf(fig)


def _write_tab6(ws, df_fixed: pd.DataFrame, period: str):
    ws.cell(row=1, column=1, value="PAINEL VISUAL — DESPESAS FIXAS").font = BODY_BOLD
    ws.cell(row=2, column=1, value=period).font = SMALL_FONT

    for cl in "ABCDEFGHIJKLMNOPQRSTUVWX":
        ws.column_dimensions[cl].width = 10

    def _embed(buf: io.BytesIO, anchor: str, w_cm: float, h_cm: float):
        buf.seek(0)
        img = XLImage(buf)
        img.width  = int(w_cm * 37.795)
        img.height = int(h_cm * 37.795)
        img.anchor = anchor
        ws.add_image(img)

    try:
        _embed(_pie_categories(df_fixed),  "A4",  13, 9)
        _embed(_bar_locations(df_fixed),   "M4",  13, 8)
        _embed(_line_trend(df_fixed),      "A30", 13, 8)
        _embed(_bar_stacked(df_fixed),     "M30", 14, 9)
        _embed(_grouped_bar(df_fixed),     "A55", 26, 12)
    except Exception as e:
        ws.cell(row=4, column=1, value=f"Erro ao gerar gráficos: {e}").font = BODY_FONT


# ---- Main entry point ----

def build_excel(df_fixed_raw: pd.DataFrame, df_excluded: pd.DataFrame,
                start: date, end: date) -> bytes:
    df_fixed = df_fixed_raw.copy()
    df_fixed["CeoCategoria"] = df_fixed.apply(assign_ceo_category, axis=1)

    df_excl = df_excluded.copy()
    if not df_excl.empty:
        df_excl["CeoCategoria"] = df_excl.apply(assign_ceo_category, axis=1)

    wb      = Workbook()
    period  = f"Período: {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"

    ws1 = wb.active; ws1.title = "Resumo Executivo"
    _write_tab1(ws1, df_fixed, start, end)

    ws2 = wb.create_sheet("Evolução Mensal")
    _write_tab2(ws2, df_fixed, period)

    ws3 = wb.create_sheet("Comparativo Localidades")
    _write_tab3(ws3, df_fixed)

    ws4 = wb.create_sheet("Despesas por Categoria")
    _write_tab4(ws4, df_fixed, period)

    ws5 = wb.create_sheet("Anomalias")
    _write_tab5(ws5, df_fixed, df_excl, period)

    ws6 = wb.create_sheet("Painel Visual")
    _write_tab6(ws6, df_fixed, period)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
